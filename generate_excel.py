import openpyxl

# Create a new workbook and a worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"

# Data to be written into the Excel file
headers = ["ID", "Name", "Age", "Email"]
data = [
    [1, "Alice", 25, "alice@example.com"],
    [2, "Bob", 30, "bob@example.com"],
    [3, "Charlie", 35, "charlie@example.com"]
]

# Write the header row
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write data rows
for row_num, row in enumerate(data, 2):
    for col_num, value in enumerate(row, 1):
        sheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to a file
file_name = "example.xlsx"
workbook.save(file_name)
print(f"Excel file '{file_name}' created successfully.")
