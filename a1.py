import random
import openpyxl
from datetime import datetime

def generate_data(num_rows):
    """
    Generate sample data for the Excel sheet.
    """
    data = []
    for i in range(1, num_rows + 1):
        entry = {
            "ID": i,
            "Name": f"User_{i}",
            "Age": random.randint(18, 60),
            "Score": random.uniform(50, 100),
            "Date": datetime.now().strftime("%Y-%m-%d")
        }
        data.append(entry)
    return data

def save_to_excel(data, filename):
    """
    Save the generated data to an Excel file.
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Data"

    # Write headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, entry in enumerate(data, start=2):
        for col_num, (key, value) in enumerate(entry.items(), start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to the specified file
    workbook.save(filename)
    print(f"Data saved to {filename}")

if __name__ == "__main__":
    # Number of rows of data to generate
    num_rows = 100

    # Generate data
    print("Generating sample data...")
    sample_data = generate_data(num_rows)

    # Save data to Excel file
    excel_filename = "generated_data.xlsx"
    print(f"Saving data to {excel_filename}...")
    save_to_excel(sample_data, excel_filename)
